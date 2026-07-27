# -*- coding: utf-8 -*-
"""ECL Workbook v4 (per-installment aging) generator.

Rebuilds the CPA "Bug Fix #7" workbook with per-installment ECL numbers.
Every calculated cell is an Excel formula; blue cells are inputs pinned to
system test goldens (source annotated in-sheet). Regenerate with:

    python docs/accounting/generate-ecl-workbook-v4.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = 'Tahoma'
F = lambda **kw: Font(name=FONT, **{'size': 10, **kw})
BLUE = Font(name=FONT, size=10, color='0000FF')          # hardcoded inputs
BLACK = Font(name=FONT, size=10)                          # formulas
GREEN = Font(name=FONT, size=10, color='008000')          # cross-sheet links
HDR = Font(name=FONT, size=10, bold=True, color='FFFFFF')
TITLE = Font(name=FONT, size=14, bold=True)
SUB = Font(name=FONT, size=9, italic=True, color='666666')
FILL_HDR = PatternFill('solid', start_color='1F4E5F')
FILL_SUB = PatternFill('solid', start_color='D9E2E8')
FILL_WARN = PatternFill('solid', start_color='FFF2CC')
FILL_OK = PatternFill('solid', start_color='E2EFDA')
FILL_FIX = PatternFill('solid', start_color='FCE4EC')
NUM = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%'
THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)

# ---- fixture (17k/12) --------------------------------------------------
PMT = 1515.83      # installment incl VAT
GROSS_PI = 1416.66  # ROUND_DOWN(17000/12)
VAT_PI = 99.17      # ROUND_HALF_UP(1190/12)
INT_PI = 500.00
PRINCIPAL = 17000.00
VAT_TOTAL = 1190.00
INTEREST = 6000.00
REPO_VALUE = 5000.00

BUCKETS = [  # label, days, rate
    ('B0 ปกติ', '0', 0.00), ('B1 1-30', '1-30', 0.02), ('B2 31-60', '31-60', 0.15),
    ('B3 61-90', '61-90', 0.50), ('B4 91-180', '91-180', 0.75), ('B5 180+', '>180', 1.00),
]

# contract: (code, name, status, paid, accrued, unpaid, oldest-age-days list)
def ages(n):
    return [15 + 30 * k for k in range(n)][::-1]  # oldest first: 195..15

PORTFOLIO = [
    ('FIN-2024-003', 'สมศักดิ์',   'ACTIVE',      3, 4,  1),
    ('FIN-2024-012', 'วิภาดา',     'ACTIVE',      5, 6,  1),
    ('FIN-2024-004', 'ประเสริฐ',   'OVERDUE',     5, 8,  3),
    ('FIN-2024-010', 'จันทร์เพ็ญ', 'TERMINATED',  5, 10, 5),
    ('FIN-2024-067', 'อนุชา',      'TERMINATED',  5, 10, 5),
    ('FIN-2024-089', 'รัตนา',      'TERMINATED',  5, 11, 6),
    ('FIN-2024-005', 'สมหมาย',     'TERMINATED',  5, 12, 7),
    ('FIN-2024-045', 'กิตติพงษ์',  'TERMINATED',  5, 12, 7),
    ('FIN-2024-023', 'ดวงใจ',      'TERMINATED',  5, 12, 7),
    ('FIN-2024-101', 'ศิริพร',     'ACTIVE (B0)', 0, 0,  0),
]

wb = Workbook()

def sheet(title, cols):
    ws = wb.create_sheet(title)
    for c, w in cols.items():
        ws.column_dimensions[c].width = w
    return ws

def put(ws, cell, val, font=BLACK, fmt=None, fill=None, align=None, border=False):
    c = ws[cell]
    c.value = val
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align: c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if border: c.border = THIN
    return c

def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HDR; c.fill = FILL_HDR; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def bucket_formula(age_cell):
    return (f'=IF({age_cell}=0,"B0 ปกติ",IF({age_cell}<=30,"B1 1-30",IF({age_cell}<=60,"B2 31-60",'
            f'IF({age_cell}<=90,"B3 61-90",IF({age_cell}<=180,"B4 91-180","B5 180+")))))')

RATE_TBL = "'3-Aging-Bucket'!$B$5:$D$10"  # label,days,rate

# ======================================================================
# Sheet 3 first (rate table referenced by others)
# ======================================================================
ws3 = sheet('3-Aging-Bucket', {'A': 3, 'B': 14, 'C': 10, 'D': 9, 'E': 12, 'F': 16, 'G': 16, 'H': 10, 'I': 16, 'J': 16})
put(ws3, 'B2', 'Aging Buckets — อัตราตั้งค่าเผื่อ (ECL v4 แยกงวด)', TITLE)
put(ws3, 'B3', 'Source: SystemConfig bad_debt_provision_rates + bad-debt.service.ts (TFRS for NPAEs Ch.13)', SUB)
header_row(ws3, 4, ['Bucket', 'วันค้าง', 'Rate'], 2)
for i, (label, days, rate) in enumerate(BUCKETS):
    r = 5 + i
    put(ws3, f'B{r}', label, BLACK, border=True)
    put(ws3, f'C{r}', days, BLACK, align='center', border=True)
    put(ws3, f'D{r}', rate, BLUE, PCT, border=True)

put(ws3, 'F2', 'สรุปรายงวดค้างทั้งพอร์ต (aggregate จาก Sheet 2)', F(bold=True))
header_row(ws3, 4, ['Bucket', 'จำนวนงวด', 'ฐานยอดค้าง (฿)', 'ค่าเผื่อ (฿)', '% ของฐาน'], 6)
for i, (label, _, _) in enumerate(BUCKETS[1:]):  # B1..B5 (B0 base 0)
    r = 5 + i
    put(ws3, f'F{r}', label, BLACK, border=True)
    put(ws3, f'G{r}', f"=COUNTIF('2-พอร์ตสัญญา'!$E$22:$E$63,F{r})", BLACK, '#,##0', border=True)
    put(ws3, f'H{r}', f"=SUMIF('2-พอร์ตสัญญา'!$E$22:$E$63,F{r},'2-พอร์ตสัญญา'!$G$22:$G$63)", BLACK, NUM, border=True)
    put(ws3, f'I{r}', f"=SUMIF('2-พอร์ตสัญญา'!$E$22:$E$63,F{r},'2-พอร์ตสัญญา'!$H$22:$H$63)", BLACK, NUM, border=True)
    put(ws3, f'J{r}', f'=IF($H$10=0,0,H{r}/$H$10)', BLACK, PCT, border=True)
put(ws3, 'F10', 'รวม', F(bold=True), fill=FILL_SUB, border=True)
for col in 'GHI':
    put(ws3, f'{col}10', f'=SUM({col}5:{col}9)', F(bold=True), NUM if col != 'G' else '#,##0', fill=FILL_SUB, border=True)
put(ws3, 'J10', '=IF(H10=0,0,H10/H10)', F(bold=True), PCT, fill=FILL_SUB, border=True)
put(ws3, 'F12', 'เช็ค: ฐานรวมต้องเท่า 11-2103 รวมพอร์ต 63,664.86 =', SUB)
put(ws3, 'I12', "=IF(ROUND(H10,2)=ROUND('2-พอร์ตสัญญา'!F17,2),\"✔ ตรง\",\"✘ ไม่ตรง\")", GREEN)
put(ws3, 'F14', 'หมายเหตุ: จำนวนงวด = นับเป็น "งวด" ไม่ใช่สัญญา (สัญญาเดียวกระจายได้หลาย bucket — วิธีแยกงวด)', SUB)

# ======================================================================
# Sheet 2 — portfolio
# ======================================================================
ws2 = sheet('2-พอร์ตสัญญา', {'A': 3, 'B': 15, 'C': 12, 'D': 13, 'E': 11, 'F': 9, 'G': 13, 'H': 13, 'I': 13, 'J': 13, 'K': 13, 'L': 30})
put(ws2, 'B2', 'พอร์ตสัญญา 10 ตัว — ค่าเผื่อหนี้สงสัยจะสูญแบบแยกงวด (ECL v4)', TITLE)
put(ws2, 'B3', 'ยอดค้างต่องวด = 1,515.83 ฿ (fixture 17,000/12 + VAT) | อายุงวดชุด 15+30k วัน | ชื่อลูกค้าสมมุติเพื่ออธิบาย', SUB)

header_row(ws2, 5, ['สัญญา', 'ลูกค้า', 'สถานะ', 'งวดค้าง', '11-2103 (฿)', 'Bucket เก่าสุด', 'ค่าเผื่อวิธีเดิม v3 (฿)', 'ค่าเผื่อแยกงวด v4 (฿)', 'ผลต่าง (release) (฿)'], 2)
mrow = 22  # matrix start
matrix_rows = []
r = 6
for code, name, status, paid, accrued, unpaid in PORTFOLIO:
    first, last = mrow, mrow + max(unpaid, 1) - 1
    put(ws2, f'B{r}', code, BLACK, border=True)
    put(ws2, f'C{r}', name, BLACK, border=True)
    put(ws2, f'D{r}', status, BLACK, border=True)
    put(ws2, f'E{r}', unpaid, BLUE, '#,##0', align='center', border=True)
    put(ws2, f'F{r}', f'=E{r}*$G$18', BLACK, NUM, border=True)
    if unpaid:
        put(ws2, f'G{r}', f'=E{mrow}', GREEN, align='center', border=True)
        put(ws2, f'H{r}', f'=ROUND(F{r}*VLOOKUP(G{r},{RATE_TBL},3,0),2)', BLACK, NUM, border=True)
        put(ws2, f'I{r}', f'=SUM(H{first}:H{last})', BLACK, NUM, border=True)
    else:
        put(ws2, f'G{r}', 'B0 ปกติ', BLACK, align='center', border=True)
        put(ws2, f'H{r}', 0, BLACK, NUM, border=True)
        put(ws2, f'I{r}', 0, BLACK, NUM, border=True)
    put(ws2, f'J{r}', f'=H{r}-I{r}', BLACK, NUM, border=True)
    for k, age in enumerate(ages(unpaid)):
        rr = mrow + k
        put(ws2, f'B{rr}', code, GREEN, border=True)
        put(ws2, f'C{rr}', accrued - unpaid + 1 + k, BLACK, '#,##0', align='center', border=True)  # installment no
        put(ws2, f'D{rr}', age, BLUE, '#,##0', align='center', border=True)
        put(ws2, f'E{rr}', bucket_formula(f'D{rr}'), BLACK, align='center', border=True)
        put(ws2, f'F{rr}', f'=VLOOKUP(E{rr},{RATE_TBL},3,0)', GREEN, PCT, border=True)
        put(ws2, f'G{rr}', '=$G$18', GREEN, NUM, border=True)
        put(ws2, f'H{rr}', f'=ROUND(G{rr}*F{rr},2)', BLACK, NUM, border=True)
        matrix_rows.append(rr)
    mrow += max(unpaid, 1) if unpaid else 0
    r += 1

put(ws2, 'B17', 'รวมพอร์ต', F(bold=True), fill=FILL_SUB, border=True)
put(ws2, 'E17', '=SUM(E6:E15)', F(bold=True), '#,##0', fill=FILL_SUB, align='center', border=True)
for col in 'FHIJ':
    put(ws2, f'{col}17', f'=SUM({col}6:{col}15)', F(bold=True), NUM, fill=FILL_SUB, border=True)
put(ws2, 'G17', '', F(bold=True), fill=FILL_SUB, border=True)
put(ws2, 'B18', 'ยอดต่องวด (฿)', SUB)
put(ws2, 'G18', PMT, BLUE, NUM)
put(ws2, 'H18', '← Source: 17,000/12 ROUND_DOWN + 1,190/12 HALF_UP (accounting.md rounding modes)', SUB)

last_matrix = matrix_rows[-1]
put(ws2, 'B20', f'รายละเอียดรายงวดค้าง (42 งวด) — แต่ละงวดใช้ rate ตาม bucket ของอายุตัวเอง', F(bold=True))
header_row(ws2, 21, ['สัญญา', 'งวดที่', 'อายุค้าง (วัน)', 'Bucket', 'Rate', 'ยอดค้าง (฿)', 'ค่าเผื่อ (฿)'], 2)
put(ws2, f'B{last_matrix + 2}', 'รวม 42 งวด', F(bold=True), fill=FILL_SUB, border=True)
put(ws2, f'G{last_matrix + 2}', f'=SUM(G22:G{last_matrix})', F(bold=True), NUM, fill=FILL_SUB, border=True)
put(ws2, f'H{last_matrix + 2}', f'=SUM(H22:H{last_matrix})', F(bold=True), NUM, fill=FILL_SUB, border=True)
nr = last_matrix + 4
put(ws2, f'B{nr}', 'หมายเหตุ TERMINATED (ACCRUED gate): ตั้งค่าเผื่อเฉพาะงวดที่ accrue เข้า 11-2103 แล้วเท่านั้น — งวดหลังบอกเลิก (2A หยุด) ไม่ตั้งสำรอง', SUB)
put(ws2, f'B{nr+1}', 'ตัวอย่างจากระบบ: TERMINATED มี 3 งวด accrued (100/70/40 วัน) + 2 งวดเลยกำหนดแต่ไม่เคย accrue → ค่าเผื่อ = 2,122.16 (ไม่ใช่ 2,183.12) — ecl-terminated-base.spec.ts', SUB)

# ======================================================================
# Sheet 4 — Scenarios สมศักดิ์
# ======================================================================
ws4 = sheet('4-Scenarios-สมศักดิ์', {'A': 3, 'B': 34, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 42})
put(ws4, 'B2', 'Scenario A-D — สมศักดิ์ FIN-2024-003 บอกเลิก+ยึด ณ งวดต่างกัน (จำลอง)', TITLE)
put(ws4, 'B3', 'จ่ายงวด 1-3 แล้วหยุด | 2A accrue ถึงงวดที่บอกเลิก | ยึดได้ราคากลาง 5,000 ฿ | ตัวเลข GL-based (ระบบ canonical — ต่างจาก count-based เดิม ±สตางค์)', SUB)

put(ws4, 'C5', 'A', HDR, fill=FILL_HDR, align='center'); put(ws4, 'D5', 'B', HDR, fill=FILL_HDR, align='center')
put(ws4, 'E5', 'C', HDR, fill=FILL_HDR, align='center'); put(ws4, 'F5', 'D', HDR, fill=FILL_HDR, align='center')
rows4 = [
    ('บอกเลิกที่งวด (k = งวดที่ accrue แล้ว)', [4, 5, 6, 7], BLUE, '#,##0'),
    ('อายุงวดค้างเก่าสุด ณ วันยึด (วัน)', [45, 75, 105, 135], BLUE, '#,##0'),
    ('งวดค้าง (k−3)', ['=C6-3', '=D6-3', '=E6-3', '=F6-3'], BLACK, '#,##0'),
]
r = 6
for label, vals, font, fmt in rows4:
    put(ws4, f'B{r}', label, BLACK, border=True)
    for i, v in enumerate(vals):
        put(ws4, f'{get_column_letter(3+i)}{r}', v, font, fmt, align='center', border=True)
    r += 1
gl_rows = [
    ('11-2103 ลูกหนี้ค้าง = (k−3)×1,515.83', "={c}8*'2-พอร์ตสัญญา'!$G$18"),
    ('11-2101 Gross คงเหลือ = 17,000 − k×1,416.66', '=17000-{c}6*1416.66'),
    ('11-2106 ดอกเบี้ยรอตัด = 6,000 − k×500', '=6000-{c}6*500'),
    ('11-2105 = 21-2102 = 1,190 − k×99.17', '=1190-{c}6*99.17'),
    ('CN VAT ม.82/5 (Dr 21-2101) = (k−3)×99.17', '={c}8*99.17'),
    ('เงินสดรับจากยึด', None),
]
for label, tmpl in gl_rows:
    put(ws4, f'B{r}', label, BLACK, border=True)
    for i in range(4):
        c = get_column_letter(3 + i)
        v = REPO_VALUE if tmpl is None else tmpl.format(c=c)
        put(ws4, f'{c}{r}', v, BLUE if tmpl is None else BLACK, NUM, border=True)
    r += 1
calc_rows = [
    ('Gross Loss = 2103+2101+2105 − เงินสด − CN VAT', '={c}9+{c}10+{c}12-{c}14-{c}13', FILL_WARN),
    ('ค่าเผื่อสะสม (แยกงวด) ณ สิ้นเดือนก่อนยึด', None, FILL_OK),
    ('Consume ค่าเผื่อ (Dr 11-2102) = min(Loss, ค่าเผื่อ)', '=MIN({c}15,{c}16)', None),
    ('Release ค่าเผื่อคงเหลือ (Dr 11-2102/Cr 51-1103)', '={c}16-{c}17', None),
    ('Net Loss 51-1102 (plug) = Loss − Consume', '={c}15-{c}17', FILL_WARN),
    ('11-2102 หลัง JP5 (ต้องเป็น 0)', '={c}16-{c}17-{c}18', FILL_OK),
    ('Net VAT นำส่งสะสมทั้งสัญญา = 1,190 − CN VAT', '=1190-{c}13', None),
    ('Net Cash Flow = 4,547.49 + 5,000 − 11,000 − NetVAT', '=4547.49+{c}14-11000-{c}21', None),
]
prov_series = [30.32, 257.69, 1015.61, 2152.48]
for label, tmpl, fill in calc_rows:
    put(ws4, f'B{r}', label, F(bold=(fill is not None)), fill=fill, border=True)
    for i in range(4):
        c = get_column_letter(3 + i)
        if tmpl is None:
            put(ws4, f'{c}{r}', prov_series[i], BLUE, NUM, fill=fill, border=True)
        else:
            put(ws4, f'{c}{r}', tmpl.format(c=c), F(bold=(fill is not None)), NUM, fill=fill, border=True)
    r += 1
put(ws4, 'G6', 'Scenario A ยึดที่ 45 วัน = ต่ำกว่าเกณฑ์ policy 60 วัน (แสดงไว้เพื่อเปรียบเทียบ)', SUB)
put(ws4, 'G7', 'อายุแถวนี้ = ณ วันยึด; ค่าเผื่อแถว 16 = snapshot cron สิ้นเดือนก่อนยึด (อายุน้อยกว่า ~30 วัน: 30/60/90/120) — สองแถวไม่ขัดกัน', SUB)
put(ws4, 'G9', 'GL หลัง 2A×k + 2B×3 — สมการเดียวกับ jp5-vat-split.spec.ts', SUB)
put(ws4, 'G15', 'A = 8,543.34 → plug 8,513.02 (ปักใน jp5-vat-split.spec.ts)', SUB)
put(ws4, 'G16', 'ชุดค่าเผื่อ ณ อายุ {30}/{60,30}/{90,60,30}/{120,90,60,30} วัน — goldens bad-debt.service.spec.ts', SUB)
put(ws4, 'G21', 'label เดิม "(Period)" → "สะสม" (แก้ตาม audit) | v3 count-based: D = 793.36, GL-based = 793.32 (เศษ 12×99.17=1,190.04)', SUB)
put(ws4, 'G22', 'v3 (แก้เลขผิดแล้ว) D = −2,245.87 count-based → GL-based = −2,245.83', SUB)

r += 2
put(ws4, f'B{r}', 'เคสจ่ายบางส่วน — pro-rate (กติกา CPA 2026-07-26)', F(bold=True, size=12)); r += 1
header_row(ws4, r, ['กรณี (งวดเดียว ค้าง 1,515.83)', 'เงินรับ', 'ค่าปรับ (FEE-FIRST)', 'ยอดค้างสุทธิ', 'ECL 31-60 (15%)', 'CN VAT pro-rate', 'CN ก่อน VAT'], 2); r += 1
pr = r
put(ws4, f'B{pr}', 'จ่าย 1,000 ไม่มีค่าปรับ', BLACK, border=True)
put(ws4, f'C{pr}', 1000, BLUE, NUM, border=True); put(ws4, f'D{pr}', 0, BLUE, NUM, border=True)
put(ws4, f'E{pr}', f"='2-พอร์ตสัญญา'!$G$18-(C{pr}-D{pr})", BLACK, NUM, border=True)
put(ws4, f'F{pr}', f'=ROUND(E{pr}*0.15,2)', BLACK, NUM, border=True)
put(ws4, f'G{pr}', f"=ROUND(99.17*E{pr}/'2-พอร์ตสัญญา'!$G$18,2)", BLACK, NUM, border=True)
put(ws4, f'H{pr}', f'=E{pr}-G{pr}', BLACK, NUM, border=True)
pr += 1
put(ws4, f'B{pr}', 'จ่าย 1,000 มีค่าปรับ 75.79 (หักค่าปรับก่อน)', BLACK, border=True)
put(ws4, f'C{pr}', 1000, BLUE, NUM, border=True); put(ws4, f'D{pr}', 75.79, BLUE, NUM, border=True)
put(ws4, f'E{pr}', f"='2-พอร์ตสัญญา'!$G$18-(C{pr}-D{pr})", BLACK, NUM, border=True)
put(ws4, f'F{pr}', f'=ROUND(E{pr}*0.15,2)', BLACK, NUM, border=True)
put(ws4, f'G{pr}', f"=ROUND(99.17*E{pr}/'2-พอร์ตสัญญา'!$G$18,2)", BLACK, NUM, border=True)
put(ws4, f'H{pr}', f'=E{pr}-G{pr}', BLACK, NUM, border=True)
put(ws4, f'B{pr+2}', 'Goldens ระบบ: ค้าง 515.83 → CN VAT 33.75 / ก่อน VAT 482.08 / ECL 77.37 | มีค่าปรับ: ค้าง 591.62 → CN VAT 38.71 (compute-cn-breakdown.spec.ts) | ใบลดหนี้จริงแสดง amountBeforeVat ตามนี้', SUB)

# ======================================================================
# Sheet 5 — JE สมศักดิ์ (Scenario A จำลอง)
# ======================================================================
ws5 = sheet('5-JE-สมศักดิ์', {'A': 3, 'B': 11, 'C': 40, 'D': 14, 'E': 14, 'F': 46})
put(ws5, 'B2', 'JE ครบชุด — สมศักดิ์ FIN-2024-003 (Scenario A จำลอง — ไม่ใช่รายการจริงใน GL)', TITLE)
put(ws5, 'B3', 'Source: case-1-overpay.csv (1A/2A/2B/จุด3) + jp5-vat-split.spec.ts Scenario A (JP5 GL-based)', SUB)

def je_block(ws, r, title, lines, note=''):
    put(ws, f'B{r}', title, F(bold=True), fill=FILL_SUB)
    if note: put(ws, f'F{r}', note, SUB)
    r += 1
    header_row(ws, r, ['บัญชี', 'ชื่อบัญชี', 'Dr (฿)', 'Cr (฿)'], 2)
    r += 1
    first = r
    for code, nm, dr, cr in lines:
        put(ws, f'B{r}', code, BLACK, border=True)
        put(ws, f'C{r}', nm, BLACK, border=True)
        put(ws, f'D{r}', dr if dr is not None else '', BLUE, NUM, border=True)
        put(ws, f'E{r}', cr if cr is not None else '', BLUE, NUM, border=True)
        r += 1
    put(ws, f'C{r}', 'รวม (ต้อง Dr = Cr)', F(bold=True), fill=FILL_SUB, border=True)
    put(ws, f'D{r}', f'=SUM(D{first}:D{r-1})', F(bold=True), NUM, fill=FILL_SUB, border=True)
    put(ws, f'E{r}', f'=SUM(E{first}:E{r-1})', F(bold=True), NUM, fill=FILL_SUB, border=True)
    put(ws, f'F{r}', f'=IF(ROUND(D{r},2)=ROUND(E{r},2),"✔ BALANCED","✘ UNBALANCED")', GREEN, border=True)
    return r + 2

r = 5
r = je_block(ws5, r, '1A — วันทำสัญญา (Dr=Cr 18,190.00)', [
    ('11-2101', 'ลูกหนี้ผ่อนชำระ Gross (ไม่รวม VAT)', PRINCIPAL, None),
    ('11-2105', 'ลูกหนี้ภาษีขายรอเรียกเก็บ', VAT_TOTAL, None),
    ('21-1101', 'เจ้าหนี้-หน้าร้าน (ยอดจัด)', None, 10000.00),
    ('21-1102', 'เจ้าหนี้ค่าคอม-หน้าร้าน', None, 1000.00),
    ('11-2106', 'รายได้รอตัดบัญชี-ดอกเบี้ย (หัก)', None, INTEREST),
    ('21-2102', 'ภาษีขายรอเรียกเก็บ', None, VAT_TOTAL),
])
r = je_block(ws5, r, 'จุดที่ 3 — จ่ายหน้าร้าน 11,000.00', [
    ('21-1101', 'เจ้าหนี้-หน้าร้าน', 10000.00, None),
    ('21-1102', 'เจ้าหนี้ค่าคอม-หน้าร้าน', 1000.00, None),
    ('11-1201', 'ธนาคาร KBank', None, 11000.00),
])
r = je_block(ws5, r, '2A — ถึงกำหนดงวด (ต่องวด, ระบบ post อัตโนมัติ 00:01 — เกิด 4 ครั้ง งวด 1-4)', [
    ('11-2103', 'ลูกหนี้ค้างชำระ (Accrual)', PMT, None),
    ('21-2102', 'ล้างภาษีขายรอเรียกเก็บ', VAT_PI, None),
    ('11-2106', 'รายได้รอตัดบัญชี-ดอกเบี้ย (ตัดงวดนี้)', INT_PI, None),
    ('11-2101', 'ลูกหนี้ Gross (ลด)', None, GROSS_PI),
    ('11-2105', 'ลูกหนี้ภาษีขายรอฯ (ล้าง)', None, VAT_PI),
    ('21-2101', 'ภาษีขาย ภ.พ.30', None, VAT_PI),
    ('41-1101', 'รายได้ดอกเบี้ย', None, INT_PI),
], note='รวมต่อครั้ง 2,115.00 × 4 ครั้ง')
r = je_block(ws5, r, '2B — รับชำระงวด (เกิด 3 ครั้ง งวด 1-3)', [
    ('11-1101', 'เงินสด', PMT, None),
    ('11-2103', 'ลูกหนี้ค้างชำระ (ล้าง)', None, PMT),
], note='× 3 ครั้ง = 4,547.49')
r = je_block(ws5, r, 'PROV — ตั้งค่าเผื่อ (cron 00:30 — งวด 4 ค้าง 1-30 วัน @2%)', [
    ('51-1103', 'ค่าเผื่อหนี้สงสัยจะสูญ (เพิ่มในปี)', 30.32, None),
    ('11-2102', 'ค่าเผื่อหนี้สงสัยจะสูญ (Contra)', None, 30.32),
])
r = je_block(ws5, r, 'JP5 — ยึดสินค้า (GL-based ทุกบรรทัด — Dr=Cr 18,435.83)', [
    ('11-1101', 'เงินสด (ราคากลางยึด)', 5000.00, None),
    ('21-2101', 'CN VAT ม.82/5 (กลับภาษีขายงวดค้าง)', 99.17, None),
    ('11-2106', 'ดอกเบี้ยรอตัดคงเหลือ (กลับ)', 4000.00, None),
    ('21-2102', 'ภาษีขายรอเรียกเก็บคงเหลือ (ล้าง)', 793.32, None),
    ('11-2102', 'Consume ค่าเผื่อ = min(Loss 8,543.34, 30.32)', 30.32, None),
    ('51-1102', 'ขาดทุนจากยึดเครื่อง (plug)', 8513.02, None),
    ('11-2103', 'ลูกหนี้ค้างชำระ (GL หลังรับ 3 งวดจริง)', None, PMT),
    ('11-2101', 'ลูกหนี้ Gross คงเหลือ (17,000−4×1,416.66)', None, 11333.36),
    ('11-2105', 'ลูกหนี้ภาษีขายรอฯ คงเหลือ', None, 793.32),
    ('21-2101', 'ภาษีขายรอเรียกเก็บถึงกำหนด (mirror 21-2102)', None, 793.32),
    ('41-1101', 'รายได้ดอกเบี้ย (mirror 11-2106 — รับรู้ตอนยึด)', None, 4000.00),
], note='Release residual = 0.00 (ค่าเผื่อ 30.32 < Loss → consume หมด) | metadata.releasedProvision = "0.00"')
put(ws5, f'B{r}', 'ถ้าค่าเผื่อ > Loss: ระบบเพิ่มบรรทัด Release (Dr 11-2102 ส่วนเหลือ / Cr 51-1103) ใน JE เดียวกัน → 11-2102 ของสัญญาเป็น 0 เสมอหลัง JP5/write-off', SUB)
put(ws5, f'B{r+1}', 'ตัวอย่างจากระบบ (write-off): ค่าเผื่อ 20,000 vs Loss 18,190 → Consume 18,190 + Release 1,810 และไม่มีบรรทัด 51-1102 เลย (bad-debt-writeoff.template.spec.ts)', SUB)

# ======================================================================
# Sheet 6 — JP5 Breakdown 6 สัญญา
# ======================================================================
ws6 = sheet('6-JP5-Breakdown', {'A': 3, 'B': 15, 'C': 11, 'D': 8, 'E': 8, 'F': 8, 'G': 12, 'H': 12, 'I': 11, 'J': 11, 'K': 11, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 11})
put(ws6, 'B2', 'JP5 Breakdown — 6 สัญญาพร้อมยึด (GL-based + Release Residual)', TITLE)
put(ws6, 'B3', 'สมมุติยึดได้ราคากลาง 5,000 ฿/เครื่อง | ค่าเผื่อ ณ วันยึด = ค่าเผื่อแยกงวด v4 | Loss = 2103+2101+2105 − เงินสด − CN VAT', SUB)
header_row(ws6, 5, ['สัญญา', 'สถานะ', 'จ่ายแล้ว j', 'Accrued k', 'ค้าง', '11-2103', '11-2101 คงเหลือ', '11-2106', '11-2105', 'CN VAT', 'เงินสดยึด', 'Gross Loss', 'ค่าเผื่อ v4', 'Consume', 'Release', 'Net Loss 51-1102', '11-2102 หลัง'], 2)
ready = [p for p in PORTFOLIO if p[2] == 'TERMINATED']
r = 6
for code, name, status, paid, accrued, unpaid in ready:
    put(ws6, f'B{r}', code, BLACK, border=True)
    put(ws6, f'C{r}', status, BLACK, border=True)
    put(ws6, f'D{r}', paid, BLUE, '#,##0', align='center', border=True)
    put(ws6, f'E{r}', accrued, BLUE, '#,##0', align='center', border=True)
    put(ws6, f'F{r}', f'=E{r}-D{r}', BLACK, '#,##0', align='center', border=True)
    put(ws6, f'G{r}', f"=F{r}*'2-พอร์ตสัญญา'!$G$18", BLACK, NUM, border=True)
    put(ws6, f'H{r}', f'=17000-E{r}*1416.66', BLACK, NUM, border=True)
    put(ws6, f'I{r}', f'=6000-E{r}*500', BLACK, NUM, border=True)
    put(ws6, f'J{r}', f'=1190-E{r}*99.17', BLACK, NUM, border=True)
    put(ws6, f'K{r}', f'=F{r}*99.17', BLACK, NUM, border=True)
    put(ws6, f'L{r}', REPO_VALUE, BLUE, NUM, border=True)
    put(ws6, f'M{r}', f'=G{r}+H{r}+J{r}-L{r}-K{r}', BLACK, NUM, border=True)
    prov_ref = 7 + PORTFOLIO.index((code, name, status, paid, accrued, unpaid)) - 1  # row in sheet2 summary
    put(ws6, f'N{r}', f"='2-พอร์ตสัญญา'!I{5 + 1 + PORTFOLIO.index((code, name, status, paid, accrued, unpaid))}", GREEN, NUM, border=True)
    put(ws6, f'O{r}', f'=MIN(M{r},N{r})', BLACK, NUM, border=True)
    put(ws6, f'P{r}', f'=N{r}-O{r}', BLACK, NUM, border=True)
    put(ws6, f'Q{r}', f'=M{r}-O{r}', BLACK, NUM, border=True)
    put(ws6, f'R{r}', f'=N{r}-O{r}-P{r}', BLACK, NUM, border=True)
    r += 1
tr = r
put(ws6, f'B{tr}', 'รวม', F(bold=True), fill=FILL_SUB, border=True)
for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
    put(ws6, f'{col}{tr}', f'=SUM({col}6:{col}{tr-1})', F(bold=True), NUM, fill=FILL_SUB, border=True)
put(ws6, f'B{tr+2}', 'คอลัมน์ Release (ใหม่ v4): สัญญา B5 accrue ครบ 12 งวด — ค่าเผื่อ 5,942.05 > Loss 4,916.66 → consume 4,916.66 + release 1,025.39 → 11-2102 = 0 (เดิม v3 ค่าเผื่อส่วนเกินค้างเงียบ)', SUB)
put(ws6, f'B{tr+3}', 'สัญญา k=12: 11-2101 เหลือ 0.08 และ 11-2105 เหลือ −0.04 (เศษปัด 12 งวด) — JP5 GL-based กวาดปิดอัตโนมัติเข้า Loss (นโยบายสตางค์: ระบบ canonical)', SUB)
put(ws6, f'B{tr+4}', 'v3 เดิมนับ 5 สัญญา (CN 3,173.44) — v4 นับ 6 สัญญาตาม summary "พร้อม JP5 = 6" ที่แก้แล้ว (รวม FIN-2024-010)', SUB)

# ======================================================================
# Sheet 7 — ติดตามบอกเลิก
# ======================================================================
ws7 = sheet('7-ติดตามบอกเลิก', {'A': 3, 'B': 15, 'C': 12, 'D': 13, 'E': 13, 'F': 16, 'G': 20, 'H': 42})
put(ws7, 'B2', 'ติดตามบอกเลิกสัญญา + สถานะ JP5', TITLE)
put(ws7, 'B3', 'เกณฑ์: ค้างเก่าสุด ≥ 90 วัน + TERMINATED = พร้อม JP5 (jp5_require_terminated_status = true — ปพพ. ม.386)', SUB)
header_row(ws7, 5, ['สัญญา', 'ลูกค้า', 'สถานะ', 'ค้าง (งวด)', 'ค้างเก่าสุด (วัน)', 'Action'], 2)
r = 6
for code, name, status, paid, accrued, unpaid in PORTFOLIO:
    oldest = ages(unpaid)[0] if unpaid else 0
    put(ws7, f'B{r}', code, BLACK, border=True)
    put(ws7, f'C{r}', name, BLACK, border=True)
    put(ws7, f'D{r}', status, BLACK, border=True)
    put(ws7, f'E{r}', unpaid, GREEN, '#,##0', align='center', border=True)
    put(ws7, f'F{r}', oldest, GREEN, '#,##0', align='center', border=True)
    action = ('=IF(AND(F{r}>=90,D{r}="TERMINATED"),"📦 พร้อม JP5",IF(F{r}>=60,"⚠ ออกหนังสือบอกเลิก",'
              'IF(F{r}>=45,"✉ หนังสือเตือน 45 วัน",IF(F{r}>0,"⏳ ติดตามปกติ","✔ ปกติ"))))').format(r=r)
    put(ws7, f'G{r}', action, BLACK, border=True)
    if code == 'FIN-2024-010':
        put(ws7, f'H{r}', '← แก้จาก v3: เดิมใส่ "รอติดตาม" ขัดกับ summary ตัวเอง (ค้าง 135 วัน ≥ 90 เหมือน FIN-2024-067)', F(size=9, italic=True, color='C00000'))
    r += 1
put(ws7, f'B{r+1}', 'สรุป: พร้อม JP5 =', F(bold=True))
put(ws7, f'E{r+1}', f'=COUNTIF(G6:G{r-1},"📦 พร้อม JP5")', F(bold=True), '#,##0', fill=FILL_OK)
put(ws7, f'F{r+1}', '(v3 นับ 6 แต่แถว FIN-2024-010 เขียนขัด — v4 แถวกับ summary ตรงกันแล้ว)', SUB)

# ======================================================================
# Sheet 1 — Executive Summary
# ======================================================================
ws1 = wb['Sheet']; ws1.title = '1-สรุปผู้บริหาร'
for c, w in {'A': 3, 'B': 44, 'C': 18, 'D': 18, 'E': 42}.items():
    ws1.column_dimensions[c].width = w
put(ws1, 'B2', 'ECL Workbook v4 — ค่าเผื่อหนี้สงสัยจะสูญแบบแยกงวด (Per-Installment Aging)', TITLE)
put(ws1, 'B3', 'BESTCHOICE FINANCE | TFRS for NPAEs Ch.13 | สร้างจากตัวเลขระบบจริง (test goldens) 2026-07-27 | แทน workbook "Bug Fix #7"', SUB)
put(ws1, 'B5', 'สิ่งที่เปลี่ยนจาก v3', F(bold=True, size=12))
changes = [
    ('1. วิธีคิดค่าเผื่อ: แยกงวด', 'แต่ละงวดค้างใช้ rate ตาม bucket ของอายุตัวเอง แล้วรวมเป็นค่าเผื่อของสัญญา (เดิม: ทั้งสัญญา × rate งวดเก่าสุด) — ตาม sheet Scenarios ที่ CPA ระบุ "ที่ถูกต้อง!"'),
    ('2. JP5/write-off: GL-based + Release', 'ทุกบรรทัดล้างตามยอด GL จริง (ปิดบั๊ก over-credit เคสจ่ายบางส่วน) + release ค่าเผื่อส่วนเกินหลัง consume → 11-2102 = 0 เสมอ'),
    ('3. TERMINATED: ACCRUED gate', 'ตั้งค่าเผื่อเฉพาะงวดที่ accrue เข้า 11-2103 แล้ว — งวดหลังบอกเลิก (ดอกเบี้ยยังไม่รับรู้) ไม่ตั้งสำรอง'),
    ('4. จ่ายบางส่วน: pro-rate + FEE-FIRST', 'ฐาน ECL และ CN VAT ลดตามยอดค้างจริง โดยหักค่าปรับออกจากเงินรับก่อน (ค่าปรับไม่ใช่การชำระเงินต้น)'),
    ('5. Streak floor: ปิด (dormant)', 'กติกา "ค้างติดต่อกัน N งวด → floor bucket" ปิดไว้ตาม workbook นี้ (ไม่มีช่อง streak) — เปิดกลับได้ด้วย config 1 แถวถ้า CPA สั่ง'),
    ('6. แก้เลขผิด v3 3 จุด', 'Scenario D Net VAT 792.85→793.36 (count) / Cash Flow D →(2,245.87) / FIN-2024-010 → พร้อม JP5'),
]
r = 6
for t, d in changes:
    put(ws1, f'B{r}', t, F(bold=True), fill=FILL_SUB, border=True)
    ws1.merge_cells(f'C{r}:E{r}')
    put(ws1, f'C{r}', d, BLACK, align='left', border=True)
    ws1.row_dimensions[r].height = 30
    r += 1
r += 1
put(ws1, f'B{r}', 'ผลกระทบต่อพอร์ต (10 สัญญา, ค้าง 42 งวด = 63,664.86 ฿)', F(bold=True, size=12)); r += 1
impact = [
    ('ค่าเผื่อวิธีเดิม v3 (ทั้งสัญญา × งวดเก่าสุด)', "='2-พอร์ตสัญญา'!H17"),
    ('ค่าเผื่อวิธีใหม่ v4 (แยกงวด)', "='2-พอร์ตสัญญา'!I17"),
    ('ลดลง (ระบบ release อัตโนมัติวันแรก — delta เทียบ GL)', "='2-พอร์ตสัญญา'!H17-'2-พอร์ตสัญญา'!I17"),
]
for t, f_ in impact:
    put(ws1, f'B{r}', t, BLACK, border=True)
    put(ws1, f'C{r}', f_, GREEN, NUM, border=True)
    r += 1
r += 1
put(ws1, f'B{r}', 'สถานะระบบ: implement + test ครบแล้วบน branch feat/ecl-per-installment (รอ CPA ยืนยัน 2 ข้อใน Sheet 8 ก่อน merge)', F(bold=True, color='C00000'), fill=FILL_WARN)

# ======================================================================
# Sheet 8 — CPA Notes
# ======================================================================
ws8 = sheet('8-CPA-Note', {'A': 3, 'B': 6, 'C': 52, 'D': 44})
put(ws8, 'B2', 'หมายเหตุถึง CPA + ขอคำยืนยัน', TITLE)
put(ws8, 'B4', 'ข้อสังเกต/นโยบายที่ระบบใช้ (แจ้งเพื่อทราบ)', F(bold=True, size=12))
notes = [
    ('ก', 'แยกงวด: provision = Σ ต่องวด ROUND(ยอดค้างสุทธิ × rate ตามอายุงวดนั้น, 2) — ปัดต่องวดก่อนรวม (ROUND_HALF_UP)', 'ตัวอย่าง: ค้าง 90/60/30 วัน = 757.92+227.37+30.32 = 1,015.61'),
    ('ข', 'TERMINATED ตั้งค่าเผื่อเฉพาะงวดที่ accrue เข้า 11-2103 แล้ว (งวดหลังบอกเลิกไม่ตั้ง — ดอกเบี้ย/VAT ยังไม่รับรู้)', 'ตัวอย่างระบบ: 3 งวด accrued + 2 งวดเลยกำหนดแต่ไม่ accrue → 2,122.16 (ไม่ใช่ 2,183.12)'),
    ('ค', 'ฐานยอดค้างหักค่าปรับก่อน (FEE-FIRST): เงินรับตัดค่าปรับก่อนแล้วค่อยตัดงวด — ฐาน ECL/CN = ยอดค้างสุทธิจริง', 'จ่าย 1,000 (มีค่าปรับ 75.79) → ตัดต้น 924.21 → ค้าง 591.62 → CN VAT 38.71'),
    ('ง', 'นโยบายสตางค์: ระบบ GL-based เป็น canonical — ต่างจากคูณจำนวนงวด ±0.04-0.08 (เศษปัด 99.17×12 = 1,190.04) ถูกกวาดปิดใน JP5', 'เช่น Net VAT Scenario D: count-based 793.36 / GL-based 793.32'),
    ('จ', 'JP5/write-off: consume ค่าเผื่อ = min(Loss, GL 11-2102) แล้ว release ส่วนเหลือ (Dr 11-2102 / Cr 51-1103) ใน JE เดียวกัน → 11-2102 ของสัญญาเป็น 0 เสมอ', 'Sheet 6 คอลัมน์ Release'),
]
r = 5
for no, txt, ex in notes:
    put(ws8, f'B{r}', no, F(bold=True), align='center', border=True)
    put(ws8, f'C{r}', txt, BLACK, align='left', border=True)
    put(ws8, f'D{r}', ex, SUB, align='left', border=True)
    ws8.row_dimensions[r].height = 42
    r += 1
r += 1
put(ws8, f'B{r}', 'ขอคำยืนยัน 2 ข้อ (ก่อนระบบขึ้น production)', F(bold=True, size=12, color='C00000')); r += 1
confirms = [
    ('1', 'ยืนยันวิธีคิดค่าเผื่อแบบ "แยกงวด" (per-installment aging) ตาม Sheet 2-3 แทนวิธีทั้งสัญญา × rate งวดเก่าสุด'),
    ('2', 'ยืนยันปิดกติกา "ค้างติดต่อกัน N งวด → floor bucket" (consecutive-missed) — workbook นี้ไม่ใช้ floor; ถ้าต้องการเปิดกลับ โปรดระบุตาราง N → bucket'),
]
for no, txt in confirms:
    put(ws8, f'B{r}', no, F(bold=True), align='center', border=True, fill=FILL_WARN)
    put(ws8, f'C{r}', txt, F(bold=True), align='left', border=True, fill=FILL_WARN)
    put(ws8, f'D{r}', 'ยืนยัน: ☐ ใช่   ☐ ปรับ (ระบุ): ____________', BLACK, align='left', border=True)
    ws8.row_dimensions[r].height = 42
    r += 1
put(ws8, f'C{r+1}', 'อ้างอิงระบบ: branch feat/ecl-per-installment | goldens: bad-debt.service.spec.ts / jp5-vat-split.spec.ts / ecl-terminated-base.spec.ts / compute-cn-breakdown.spec.ts', SUB)

# order sheets
order = ['1-สรุปผู้บริหาร', '2-พอร์ตสัญญา', '3-Aging-Bucket', '4-Scenarios-สมศักดิ์', '5-JE-สมศักดิ์', '6-JP5-Breakdown', '7-ติดตามบอกเลิก', '8-CPA-Note']
wb._sheets = [wb[n] for n in order]

import os
out = os.path.join(os.path.dirname(__file__), 'ECL-workbook-v4-per-installment.xlsx')
wb.save(out)
print('saved', out)
